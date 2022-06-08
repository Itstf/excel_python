from operator import index
from tqdm import tqdm
from time import sleep
import os
import sys
sys.path.append(os.path.realpath("."))
import inquirer
import pandas as pd
from openpyxl.styles import PatternFill, Font
from openpyxl import Workbook, load_workbook
import matplotlib.pyplot as plt

# ----------------------------------------------

class Op_chanceColor():
    def options_changeColor():
        os.system('cls')
        print('Olá :)')
        sleep(0.5)
        os.system('cls')
        print(''' 
                Se for atualizar a planilha, não mude a cor!
                Obrigado!
              ''')
        sleep(1.3)
        os.system('cls')
        wb = load_workbook('servidores.xlsx')
        ws = wb.active
        questions = [
            inquirer.List(
                'mudar_cor',
                message='Para qual cor deseja mudar?',
                choices=['\33[37mCinza\33[m', '\33[36mCiano\33[m', '\33[31mVermelho\33[m', 'Pink'],
            ),
        ]
        answers = inquirer.prompt(questions)['mudar_cor']
        sleep(1)
        os.system('cls')
        print('\33[33mAlterando cor. .\33[m')
        for i in tqdm (range (50), desc="\33[33mLoading...\33[m"): 
            sleep(0.01)
            pass
        sleep(1)
        os.system('cls')
        if answers == '\33[37mCinza\33[m':
            preenchimento = PatternFill(start_color='CACACA',
                    end_color='CACACA',
                    fill_type='solid')
            negrito = Font(bold=True)
            cabecalho = ['A1', 'B1', 'C1', 'D1', 'E1', 'F1', 'G1', 'H1', 'I1', 'J1', 'K1', 'L1', 'M1', 'N1', 'O1', 'P1', 'Q1', 'R1', 'S1', 'T1', 'U1', 'V1', 'W1', 'X1', 'Y1', 'Z1' , 'AA1', 'AB1', 'AC1', 'AD1', 'AE1', 'AF1', 'AG1', 'AH1', 'AI1', 'AJ1', 'AK1', 'AL1', 'AM1', 'AN1', 'AO1', 'AP1', 'AQ1', 'AR1', 'AS1', 'AT1', 'AU1', 'AV1', 'AW1', 'AX1', 'AY1', 'AZ1']
            for cell in cabecalho:
                ws[cell].fill = preenchimento
                ws[cell].font = negrito
            wb.save('servidores.xlsx')
        elif answers == '\33[36mCiano\33[m':
            preenchimento = PatternFill(start_color='00BBC9',
                    end_color='00BBC9',
                    fill_type='solid')
            negrito = Font(bold=True)
            cabecalho = ['A1', 'B1', 'C1', 'D1', 'E1', 'F1', 'G1', 'H1', 'I1', 'J1', 'K1', 'L1', 'M1', 'N1', 'O1', 'P1', 'Q1', 'R1', 'S1', 'T1', 'U1', 'V1', 'W1', 'X1', 'Y1', 'Z1' , 'AA1', 'AB1', 'AC1', 'AD1', 'AE1', 'AF1', 'AG1', 'AH1', 'AI1', 'AJ1', 'AK1', 'AL1', 'AM1', 'AN1', 'AO1', 'AP1', 'AQ1', 'AR1', 'AS1', 'AT1', 'AU1', 'AV1', 'AW1', 'AX1', 'AY1', 'AZ1']
            for cell in cabecalho:
                ws[cell].fill = preenchimento
                ws[cell].font = negrito
            wb.save('servidores.xlsx')
        elif answers == '\33[31mVermelho\33[m':
            preenchimento = PatternFill(start_color='FF4858',
                    end_color='FF4858',
                    fill_type='solid')
            negrito = Font(bold=True)
            cabecalho = ['A1', 'B1', 'C1', 'D1', 'E1', 'F1', 'G1', 'H1', 'I1', 'J1', 'K1', 'L1', 'M1', 'N1', 'O1', 'P1', 'Q1', 'R1', 'S1', 'T1', 'U1', 'V1', 'W1', 'X1', 'Y1', 'Z1' , 'AA1', 'AB1', 'AC1', 'AD1', 'AE1', 'AF1', 'AG1', 'AH1', 'AI1', 'AJ1', 'AK1', 'AL1', 'AM1', 'AN1', 'AO1', 'AP1', 'AQ1', 'AR1', 'AS1', 'AT1', 'AU1', 'AV1', 'AW1', 'AX1', 'AY1', 'AZ1']
            for cell in cabecalho:
                ws[cell].fill = preenchimento
                ws[cell].font = negrito
            wb.save('servidores.xlsx')
        elif answers == 'Pink':
            preenchimento = PatternFill(start_color='FF1493',
                    end_color='FF1493',
                    fill_type='solid')
            negrito = Font(bold=True)
            cabecalho = ['A1', 'B1', 'C1', 'D1', 'E1', 'F1', 'G1', 'H1', 'I1', 'J1', 'K1', 'L1', 'M1', 'N1', 'O1', 'P1', 'Q1', 'R1', 'S1', 'T1', 'U1', 'V1', 'W1', 'X1', 'Y1', 'Z1' , 'AA1', 'AB1', 'AC1', 'AD1', 'AE1', 'AF1', 'AG1', 'AH1', 'AI1', 'AJ1', 'AK1', 'AL1', 'AM1', 'AN1', 'AO1', 'AP1', 'AQ1', 'AR1', 'AS1', 'AT1', 'AU1', 'AV1', 'AW1', 'AX1', 'AY1', 'AZ1']
            for cell in cabecalho:
                ws[cell].fill = preenchimento
                ws[cell].font = negrito
            wb.save('servidores.xlsx')